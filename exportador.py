"""
exportador.py - Exportación de datos financieros a Excel (XLSX) y CSV.

Genera archivos listos para abrir y manipular en Excel:
- XLSX: un solo archivo con hojas Resumen, Movimientos (paginada) y Gastos por categoría.
- CSV: archivos UTF-8 (BOM) con la tabla de movimientos, paginados si superan el límite.

Sin dependencias pesadas: usa solo la biblioteca estándar y openpyxl (import perezoso).
"""

import csv
import os
import re
from datetime import date, datetime, timedelta
from typing import Optional

MAX_TRANSACCIONES = 1_000_000
MAX_FILAS_POR_HOJA = 100_000

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

_HEADERS_MOVIMIENTOS = ["#", "Fecha", "Tipo", "Categoría", "Descripción", "Monto", "Moneda"]
_HEADERS_CSV = ["Fecha", "Tipo", "Categoría", "Descripción", "Monto", "Moneda"]


# ============================================================
# PERÍODOS
# ============================================================

def _resolver_periodo(periodo: Optional[str] = None):
    """Resuelve el período a (etiqueta, fecha_inicio, fecha_fin).

    periodos: 'todo' | 'mes' | '30' | 'YYYY-MM' | None.
    Devuelve inicio/fin en formato 'YYYY-MM-DD' o None (sin filtro).
    """
    hoy = date.today()
    p = (periodo or "todo").strip().lower()

    if p in ("mes", "este mes", "mensual"):
        inicio = hoy.replace(day=1)
        return f"{MESES_ES[hoy.month]} {hoy.year}", inicio.isoformat(), hoy.isoformat()

    if p in ("30", "30d", "30 dias", "últimos 30 días", "ultimos 30 dias", "treinta"):
        inicio = hoy - timedelta(days=29)
        return "Últimos 30 días", inicio.isoformat(), hoy.isoformat()

    m = re.match(r"^(\d{4})-(\d{1,2})$", p)
    if m:
        anio, mes = int(m.group(1)), int(m.group(2))
        try:
            fin_mes = date(anio, mes + 1, 1) - timedelta(days=1) if mes < 12 else date(anio, 12, 31)
        except ValueError:
            return "Todo el historial", None, None
        return f"{MESES_ES.get(mes, str(mes))} {anio}", date(anio, mes, 1).isoformat(), fin_mes.isoformat()

    return "Todo el historial", None, None


def mapear_periodo_ia(fecha_ia: Optional[str], mensaje: str) -> str:
    """Convierte la salida libre de la IA sobre el período en un valor conocido."""
    texto = (fecha_ia or "") + " " + (mensaje or "")
    t = texto.lower()

    if re.search(r"\d{4}-\d{1,2}", t):
        m = re.search(r"(\d{4})-(\d{1,2})", t)
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    nombres_mes = {nombre.lower(): num for num, nombre in MESES_ES.items()}
    m_mes = re.search(r"\b(" + "|".join(nombres_mes.keys()) + r")\b", t)
    if m_mes:
        mes_num = nombres_mes[m_mes.group(1)]
        m_anio = re.search(r"\b(20\d{2})\b", t)
        anio = int(m_anio.group(1)) if m_anio else date.today().year
        return f"{anio}-{mes_num:02d}"
    if "este mes" in t or "mes actual" in t or "del mes" in t or re.search(r"\bmes\b", t):
        return "mes"
    if re.search(r"30\s*d[ií]as", t) or re.search(r"\b30\b", t):
        return "30"
    if "todo" in t or "historial" in t or "todos" in t:
        return "todo"
    return "todo"


def _detectar_formato(mensaje: str) -> Optional[str]:
    """Detecta el formato pedido ('csv' o 'xlsx') en el mensaje, si se menciona."""
    t = (mensaje or "").lower()
    if "csv" in t:
        return "csv"
    if "xlsx" in t or "excel" in t:
        return "xlsx"
    return None


# ============================================================
# UTILIDADES DE DATOS
# ============================================================

def _limpiar_descripcion(texto: Optional[str]) -> str:
    """Limpia la descripción cruda de una transacción para la exportación."""
    d = (texto or "").strip()
    for pref in ("gasto: ", "ingreso: "):
        if d.lower().startswith(pref):
            d = d[len(pref):].strip()
            break
    for pref in ("gasté ", "gaste ", "recibí ", "recibi ", "compré ", "compre ",
                 "pagué ", "pague ", "vendí ", "vendi "):
        if d.lower().startswith(pref):
            d = d[len(pref):].strip()
            break
    d = re.sub(r'^\$\s*\d+(?:[.,]\d+)?\s*(?:cup|usd|usdt|mlc|eur)?\s*', '', d).strip()
    for pref in ("en ", "de ", "del "):
        if d.lower().startswith(pref):
            d = d[len(pref):].strip()
            break
    return d or "Sin descripción"


def _mapa_monedas(monedas: list) -> dict:
    """Construye {moneda_id: {abreviatura, simbolo, nombre}}."""
    mapa = {}
    for m in monedas or []:
        try:
            mapa[m["id"]] = {
                "abreviatura": m.get("abreviatura", ""),
                "simbolo": m.get("simbolo", "$"),
                "nombre": m.get("nombre", ""),
            }
        except (KeyError, TypeError):
            continue
    return mapa


def _abrev_moneda(moneda_id, mapa_monedas) -> str:
    m = mapa_monedas.get(moneda_id) if moneda_id else None
    return m["abreviatura"] if m else "Sin moneda"


def _agrupar_gastos_por_categoria(transacciones: list, mapa_monedas: dict) -> dict:
    """Agrupa gastos por (categoría, moneda): { (cat, moneda): monto }."""
    agg = {}
    for t in transacciones or []:
        if t.get("tipo") != "gasto":
            continue
        cat = t.get("categoria_nombre") or "Otros"
        abrev = _abrev_moneda(t.get("moneda_id"), mapa_monedas)
        clave = (cat, abrev)
        try:
            agg[clave] = agg.get(clave, 0.0) + float(t.get("cantidad", 0))
        except (TypeError, ValueError):
            continue
    return agg


# ============================================================
# GENERACIÓN XLSX
# ============================================================

def _estilos():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    return {
        "azul": "1F4E79",
        "gris_claro": "F2F7FB",
        "verde": "107C10",
        "rojo": "B00020",
        "header_fill": PatternFill("solid", fgColor="1F4E79"),
        "zebra_fill": PatternFill("solid", fgColor="F2F7FB"),
        "header_font": Font(bold=True, color="FFFFFF", size=11),
        "titulo_font": Font(bold=True, size=16, color="1F4E79"),
        "sub_font": Font(size=11, color="555555"),
        "bold": Font(bold=True),
        "thin": Side(style="thin", color="C8D4E0"),
        "borde": Border(left=Side(style="thin", color="C8D4E0"),
                        right=Side(style="thin", color="C8D4E0"),
                        top=Side(style="thin", color="C8D4E0"),
                        bottom=Side(style="thin", color="C8D4E0")),
        "center": Alignment(horizontal="center", vertical="center"),
        "right": Alignment(horizontal="right", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
    }


def _autoancho(ws, encabezados: list, filas: list, max_cols: int = 60):
    from openpyxl.utils import get_column_letter
    for i, encabezado in enumerate(encabezados, start=1):
        ancho = len(str(encabezado))
        for fila in filas:
            valor = fila[i - 1] if i - 1 < len(fila) else None
            if valor is not None and not isinstance(valor, (int, float)):
                ancho = max(ancho, min(len(str(valor)), 40))
        ws.column_dimensions[get_column_letter(i)].width = min(ancho + 2, max_cols)


def _escribir_hoja_movimientos(wb, nombre_hoja: str, filas: list, monedas_abrev: set, est):
    from openpyxl.styles import Font
    ws = wb.create_sheet(title=nombre_hoja)
    ws.append(_HEADERS_MOVIMIENTOS)
    for c in range(1, len(_HEADERS_MOVIMIENTOS) + 1):
        celda = ws.cell(row=1, column=c)
        celda.fill = est["header_fill"]
        celda.font = est["header_font"]
        celda.alignment = est["center"]
        celda.border = est["borde"]

    for i, fila in enumerate(filas, start=2):
        ws.append(fila)
        fill = est["zebra_fill"] if i % 2 == 0 else None
        for c in range(1, len(fila) + 1):
            celda = ws.cell(row=i, column=c)
            celda.border = est["borde"]
            if c in (1, 2, 7):
                celda.alignment = est["center"]
            elif c == 6:
                celda.alignment = est["right"]
                celda.number_format = '#,##0.00'
            else:
                celda.alignment = est["left"]
            if fill:
                celda.fill = fill
            if c == 3:
                celda.font = Font(bold=True, color=est["rojo"] if fila[2] == "Gasto" else est["verde"])

    ultima_fila = ws.max_row + 2
    ws.cell(row=ultima_fila, column=1, value="TOTALES DEL PERÍODO").font = est["bold"]
    fila_total = ultima_fila + 1
    ws.cell(row=fila_total, column=3, value="Tipo").font = est["bold"]
    ws.cell(row=fila_total, column=5, value="Monto").font = est["bold"]
    ws.cell(row=fila_total, column=7, value="Moneda").font = est["bold"]
    for abrev in sorted(monedas_abrev):
        total = sum(f[5] for f in filas if f[6] == abrev)
        fila_total += 1
        ws.cell(row=fila_total, column=4, value=f"TOTAL ({abrev})").font = est["bold"]
        celda = ws.cell(row=fila_total, column=6, value=total)
        celda.font = est["bold"]
        celda.number_format = '#,##0.00'
        celda.alignment = est["right"]
        ws.cell(row=fila_total, column=7, value=abrev).font = est["bold"]

    ws.freeze_panes = "A2"
    if filas:
        ws.auto_filter.ref = f"A1:G{len(filas) + 1}"
    _autoancho(ws, _HEADERS_MOVIMIENTOS, filas)


def generar_xlsx(usuario_id: int, nombre_usuario: str, label: str,
                 balance: dict, transacciones: list, monedas: list,
                 directorio: str) -> str:
    """Genera el archivo XLSX y devuelve su ruta."""
    from openpyxl import Workbook

    est = _estilos()
    mapa_monedas = _mapa_monedas(monedas)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = os.path.join(directorio, f"finanzas_{usuario_id}_{stamp}.xlsx")

    filas_mov = []
    monedas_abrev = set()
    for i, t in enumerate(transacciones, start=1):
        fecha = (t.get("fecha") or "")[:10]
        tipo = "Gasto" if t.get("tipo") == "gasto" else "Ingreso"
        cat = t.get("categoria_nombre") or "Otros"
        abrev = _abrev_moneda(t.get("moneda_id"), mapa_monedas)
        monto = float(t.get("cantidad", 0))
        filas_mov.append([i, fecha, tipo, cat, _limpiar_descripcion(t.get("descripcion")), monto, abrev])
        monedas_abrev.add(abrev)

    paginas = [filas_mov[i:i + MAX_FILAS_POR_HOJA] for i in range(0, len(filas_mov), MAX_FILAS_POR_HOJA)]
    if not paginas:
        paginas = [[]]

    wb = Workbook()

    # ---- Hoja Resumen ----
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Resumen Financiero"
    ws["A1"].font = est["titulo_font"]
    ws["A2"] = f"Período: {label}"
    ws["A2"].font = est["sub_font"]
    ws["A3"] = f"Usuario: {nombre_usuario or usuario_id}"
    ws["A3"].font = est["sub_font"]
    ws["A4"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A4"].font = est["sub_font"]

    ws["A6"] = "Balance por moneda"
    ws["A6"].font = est["bold"]
    enc = ["Moneda", "Ingresos", "Gastos", "Neto"]
    for c, h in enumerate(enc, start=1):
        celda = ws.cell(row=7, column=c, value=h)
        celda.fill = est["header_fill"]
        celda.font = est["header_font"]
        celda.alignment = est["center"]
        celda.border = est["borde"]

    por_moneda = balance.get("por_moneda", {}) or {}
    fila = 8
    if por_moneda and not (len(por_moneda) == 1 and "Sin moneda" in por_moneda):
        for abrev, d in sorted(por_moneda.items()):
            simbolo = d.get("simbolo", "$")
            ingresos = float(d.get("ingresos", 0))
            gastos = float(d.get("gastos", 0))
            valores = [f"{simbolo} {abrev}", ingresos, gastos, ingresos - gastos]
            for c, v in enumerate(valores, start=1):
                celda = ws.cell(row=fila, column=c, value=v)
                celda.border = est["borde"]
                if c == 1:
                    celda.alignment = est["left"]
                else:
                    celda.alignment = est["right"]
                    celda.number_format = '#,##0.00'
            fila += 1
    else:
        simbolo = "$"
        ingresos = float(balance.get("ingresos", 0))
        gastos = float(balance.get("gastos", 0))
        for c, v in enumerate([f"{simbolo} Sin moneda", ingresos, gastos, ingresos - gastos], start=1):
            celda = ws.cell(row=fila, column=c, value=v)
            celda.border = est["borde"]
            celda.alignment = est["center" if c == 1 else "right"]
            if c > 1:
                celda.number_format = '#,##0.00'

    fila += 1
    ws.cell(row=fila, column=1, value=f"Transacciones exportadas: {len(filas_mov)}").font = est["sub_font"]
    if len(paginas) > 1:
        ws.cell(row=fila + 1, column=1,
                value=f"Movimientos divididos en {len(paginas)} hojas por el límite de filas.").font = est["sub_font"]

    _autoancho(ws, enc, [])

    # ---- Hojas de Movimientos (paginadas) ----
    for idx, chunk in enumerate(paginas, start=1):
        nombre = "Movimientos" if idx == 1 else f"Movimientos ({idx})"
        _escribir_hoja_movimientos(wb, nombre, chunk, monedas_abrev, est)

    # ---- Hoja Gastos por categoría ----
    agg = _agrupar_gastos_por_categoria(transacciones, mapa_monedas)
    wsg = wb.create_sheet(title="Gastos por categoría")
    encg = ["Categoría", "Moneda", "Monto", "% del período"]
    wsg.append(encg)
    for c in range(1, len(encg) + 1):
        celda = wsg.cell(row=1, column=c)
        celda.fill = est["header_fill"]
        celda.font = est["header_font"]
        celda.alignment = est["center"]
        celda.border = est["borde"]

    if agg:
        totales_moneda = {}
        for (cat, abrev), monto in agg.items():
            totales_moneda[abrev] = totales_moneda.get(abrev, 0.0) + monto
        filasg = sorted(((cat, abrev, monto, monto / totales_moneda[abrev] if totales_moneda.get(abrev) else 0)
                         for (cat, abrev), monto in agg.items()),
                        key=lambda x: (x[2], x[1]), reverse=True)
        for i, (cat, abrev, monto, pct) in enumerate(filasg, start=2):
            wsg.append([cat, abrev, monto, pct])
            for c in range(1, 5):
                celda = wsg.cell(row=i, column=c)
                celda.border = est["borde"]
                if c == 3:
                    celda.alignment = est["right"]
                    celda.number_format = '#,##0.00'
                elif c == 4:
                    celda.alignment = est["right"]
                    celda.number_format = '0.0%'
                else:
                    celda.alignment = est["left"]
            if i % 2 == 0:
                for c in range(1, 5):
                    wsg.cell(row=i, column=c).fill = est["zebra_fill"]
        wsg.freeze_panes = "A2"
        _autoancho(wsg, encg, filasg)
    else:
        wsg["A2"] = "Sin gastos en el período."

    wb.save(ruta)
    return ruta


# ============================================================
# GENERACIÓN CSV
# ============================================================

def _filas_csv(transacciones: list, mapa_monedas: dict) -> list:
    filas = []
    for t in transacciones:
        fecha = (t.get("fecha") or "")[:10]
        tipo = "Gasto" if t.get("tipo") == "gasto" else "Ingreso"
        cat = t.get("categoria_nombre") or "Otros"
        abrev = _abrev_moneda(t.get("moneda_id"), mapa_monedas)
        monto = float(t.get("cantidad", 0))
        filas.append([fecha, tipo, cat, _limpiar_descripcion(t.get("descripcion")), monto, abrev])
    return filas


def generar_csv_partes(usuario_id: int, label: str, transacciones: list, monedas: list,
                       directorio: str) -> list:
    """Genera archivos CSV (UTF-8 con BOM) y devuelve la lista de rutas."""
    mapa_monedas = _mapa_monedas(monedas)
    filas = _filas_csv(transacciones, mapa_monedas)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    paginas = [filas[i:i + MAX_FILAS_POR_HOJA] for i in range(0, len(filas), MAX_FILAS_POR_HOJA)]
    if not paginas:
        paginas = [[]]

    rutas = []
    total = len(paginas)
    for idx, chunk in enumerate(paginas, start=1):
        sufijo = "" if total == 1 else f"_parte_{idx}"
        ruta = os.path.join(directorio, f"finanzas_{usuario_id}_{stamp}{sufijo}.csv")
        with open(ruta, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(_HEADERS_CSV)
            writer.writerows(chunk)
        rutas.append(ruta)
    return rutas